from bsedata.bse import BSE
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

b = BSE()
driver = webdriver.Edge()
url = "https://www.bseindia.com/corporates/Forth_Results.html?expandable=0"
driver.get(url)
time.sleep(2)

codes = [line.rstrip() for line in open('Input.txt')]

data = []

for target_code in codes:
    flag = 0
    try:
        row = driver.find_element(By.XPATH, f"//tr[td/a[contains(text(), '{target_code}')]]")
    except:
        date = "N/A"
        flag = 1
    if flag == 0:
        company_name = row.find_elements(By.TAG_NAME, "td")[1].text
        date = row.find_elements(By.TAG_NAME, "td")[2].text
    try:
        quote = b.getQuote(target_code)
        company_name = quote['companyName']
    except:
        data.append({
            'Company Code': target_code,
            'Company Name': "None",
            'Current Value': "None",
            'high': "None",
            'Low': "None",
            'Previous Close': "None",
            '52H': "None",
            '52L': "None",
            '%from52H': "Failed",
            'Date': date
        })
        print(f"Failed to get for {target_code}")
        continue
    
    percentage_from_52H = ((float(quote['52weekHigh']) - float(quote['currentValue'])) / float(quote['52weekHigh'])) * 100
    percentage_from_52H = round(percentage_from_52H, 2)
    data.append({
        'Company Code': target_code,
        'Company Name': quote['companyName'],
        'Current Value': quote['currentValue'],
        'high': quote['dayHigh'],
        'Low': quote['dayLow'],
        'Previous Close': quote['previousClose'],
        '52H': quote['52weekHigh'],
        '52L': quote['52weekLow'],
        '%from52H': percentage_from_52H,
        'Date': date
    })
    time.sleep(2)
    print(f"Data fetched for {target_code}")

driver.quit()

df = pd.DataFrame(data)
output_file = 'output.xlsx'

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df.to_excel(writer, index=False)
    worksheet = writer.sheets['Sheet1']
    
    for column in df.columns:
        column_length = max(df[column].astype(str).map(len).max(), len(column))
        col_idx = df.columns.get_loc(column)
        worksheet.column_dimensions[get_column_letter(col_idx+1)].width = column_length
    
    for row in range(2, len(df) + 2):
        cell = worksheet.cell(row=row, column=df.columns.get_loc('%from52H') + 1)
        cell1 = worksheet.cell(row=row, column=df.columns.get_loc('Company Name') + 1)
        if cell.value is not None and cell.value != "Failed" and int(cell.value) >= 24:
            cell.fill = PatternFill(start_color="50C878", end_color="50C878", fill_type="solid")
            cell1.fill = PatternFill(start_color="50C878", end_color="50C878", fill_type="solid")

print("Data saved to", output_file)